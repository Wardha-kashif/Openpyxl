import openpyxl
from openpyxl.styles import numbers

wb=openpyxl.load_workbook("result.xlsx")
print(wb.sheetnames)
ws=wb["Sheet1"]
print(ws)

# ws['c4']='11/09/20'
ws['E4'].number_format=numbers.FORMAT_TEXT
wb.save("result.xlsx")