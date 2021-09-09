import openpyxl
from openpyxl.styles import Font,Color

wb=openpyxl.load_workbook("result.xlsx")
print(wb.sheetnames)
ws=wb["Sheet1"]
print(ws)

font_style=Font(color='1A4FDF',size=24,name="Chalkboard",italic=True)

b4=ws['B4']

b4.font=font_style


col_style=font_style=Font(color='1A4FDF',size=12,name="Chalkboard",underline='single',strikethrough=True)
for i in range(2,7):
    ws.cell(row=i,column=3).font=col_style

wb.save("result.xlsx")