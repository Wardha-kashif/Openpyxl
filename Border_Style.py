import openpyxl
from openpyxl.styles import Border,Side

wb=openpyxl.load_workbook("result.xlsx")
print(wb.sheetnames)
ws=wb["Sheet1"]
print(ws)


top=Side(border_style='dashed',color='1A4FDF')
bottom=Side(border_style='double',color='1A4FDF')
border=Border(top=top)

ws['E6'].border=border

wb.save('result.xlsx')