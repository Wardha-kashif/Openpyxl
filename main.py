import openpyxl

#Part 1
from openpyxl import utils

wb=openpyxl.load_workbook("result.xlsx")
print(wb.sheetnames)
ws=wb["Sheet1"]
print(ws)

#Creating the sheet
# wb1=wb.create_sheet("Sheet2")
#at particular position

# wb1=wb.create_sheet("Sheet3",0)
# wb.save("result.xlsx")


#Part 2

print(ws['B3'].value)
print(ws.cell(row=3,column=4).value)

value_range=ws['B2':'C5']

for a,b in value_range:
    print(a.value,b.value)

#Part 3

rows=ws.iter_rows(min_row=1,max_row=7,min_col=1,max_col=2)
print(rows)

ID=[]
names=[]

for a,b in rows:
    ID.append(a.value)
    names.append(b.value)

print(ID)
print(names)


cols=ws.iter_cols(min_row=1,max_row=7,min_col=1,max_col=7)
print(cols)
# ID1=[]
# for a,b,c,d,e,f,g in cols:
#     ID1.append(f.value)
# print(ID1)

# ID1=[]
# names1=[]
#
# for a,b in cols:
#     ID1.append(a.value)
#     names1.append(b.value)
#
# print(ID1)
# print(names1)


rows=list(ws.rows)
print(rows)

columns=list(ws.columns)
print(columns)

for col in ws.iter_cols(min_row=1, max_row=1):
    for mycell in col:
        if mycell.value == "Percentage":
            origCol = mycell.column
            print(origCol)
#     # get the column letter for the first empty column to output the new values
# newCol = utils.get_column_letter(ws.max_column+1)
#
#
# for myrow in range(2, ws.max_row+1):
#     myrow = str(myrow)
#     # do some stuff to make the new value
#     cleanedResp = (ws[origCol + myrow].value)
#     ws[newCol + myrow] = cleanedResp
#
# wb.save("result.xlsx")

ws['B5'].value="Amir"
ws['H1']="Total"

for i in range(2,7):
   b_col= ws.cell(row=i,column=7).value
   c_val=b_col *2
   ws.cell(row=i,column=8).value=c_val
wb.save("result.xlsx")