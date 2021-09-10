import openpyxl
from openpyxl.styles import Border,Side,Font

wb=openpyxl.load_workbook("result.xlsx")
print(wb.sheetnames)
ws=wb["Sheet1"]
print(ws)

ws['I2']="=SUM(H2:H3)"


ws['I3']="=AVERAGE(H2:H3)"

# (Blance * interest rate ) + Balance

# ws['J']="Balance"
# ws['J1'].font=Font(bold=True,name='Arial',size="10")
for i in range(2,7):
    balance=ws.cell(row=i,column=7).value
    # intrest = ws.cell(row=i, column=3).value
    # final_balance=(balance * interest) + balance
    # ws.cell(row=i,column=3).value=final_balance
    # print(balance)

wb.save('result.xlsx')


#Balance After Year

# (Blance * interest rate ) + Balance

# ws['J']="Balance"
# ws['J1'].font=Font(bold=True,name='Arial',size="10")