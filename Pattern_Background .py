import openpyxl
from openpyxl.styles import PatternFill

wb=openpyxl.load_workbook("result.xlsx")
print(wb.sheetnames)
ws=wb["Sheet1"]
print(ws)

fill_pattern=PatternFill(patternType='solid',fgColor="1A4FDF")

ws['E4'].fill=fill_pattern
wb.save('result.xlsx')