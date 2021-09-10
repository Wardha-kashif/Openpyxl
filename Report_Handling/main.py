import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

#Reading the Excel File
excel_file=pd.read_excel("supermarket_sales - Sheet1.xlsx")
excel_file[["Gender","Product line","Total"]]

report_table = excel_file.pivot_table(index='Gender',
                                      columns='Product line',
                                      values='Total',
                                      aggfunc='sum').round(0)


report_table.to_excel('report_2021.xlsx',
                      sheet_name='Report',
                      startrow=4)

wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']
# cell references (original spreadsheet)
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row


wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']

# barchart
barchart = BarChart()

#locate data and categories
data = Reference(sheet,
                 min_col=min_column+1,
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row) #including headers
categories = Reference(sheet,
                       min_col=min_column,
                       max_col=min_column,
                       min_row=min_row+1,
                       max_row=max_row) #not including headers
# adding data and categories
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

#location chart
sheet.add_chart(barchart, "B12")
barchart.title = 'Sales by Product line'
barchart.style = 3

#choose the chart style
wb.save('report_2021.xlsx')

# Applying Excel formulas through Python
#https://towardsdatascience.com/a-simple-guide-to-automate-your-excel-reporting-with-python-9d35f143ef7

# first, we
# need
# to
# get
# the
# alphabet
# to
# have
# it as a
# reference
# for the names that columns have in Excel (A, B, C, …) To do so, we use the string library and write the following code.
import string
alphabet = list(string.ascii_uppercase)
excel_alphabet = alphabet[0:max_column]
print(excel_alphabet)

sheet['B7'] = '=SUM(B5:B6)'
sheet['B7'].style = 'Currency'

print(min_column)
print(max_column)
print(min_row)
print(max_row)
wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']
# sum in columns B-G
for i in excel_alphabet:
    if i!='A':
        sheet[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
        sheet[f'{i}{max_row+1}'].style = 'Currency'
# adding total label
sheet[f'{excel_alphabet[0]}{max_row+1}'] = 'Total'
wb.save('report_2021.xlsx')

#Formatting the report sheet
wb = load_workbook('report_2021.xlsx')
sheet = wb['Report']
sheet['A1'] = 'Sales Report'
sheet['A2'] = '2021'
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)
wb.save('report_2021.xlsx')

